from decimal import ROUND_HALF_UP, Decimal

import click
from openpyxl import load_workbook


@click.group()
def cli():
    """
    Validate the accuracy and consistency of IS daily reports
    """
    pass


@cli.command(name="load-data")
@click.option(
    '--start_day', prompt='Enter start day', default=1, type=int, show_default=True
)
@click.option(
    '--end_day', prompt='Enter end day', default=31, type=int, show_default=True
)
@click.option(
    '--file_path', prompt='Enter file path', default='daily.xlsx', show_default=True
)
def load_is_daily(file_path, start_day, end_day):
    # اگر پارامترها داده نشده باشند، مقدار پیش‌فرض بده
    if start_day < 0:
        start_day = 0
    if end_day > 30:
        end_day = 30

    # باز کردن و خواندن اکسل گزارش روزانه در حالت خواندن برای سرعت بیشتر
    try:
        wb_daily_report = load_workbook(
            filename=rf"{file_path}", data_only=True, read_only=True
        )
    except Exception as e:
        click.secho(message=f"There is a problem with excel file: {e}", fg="red")

    full_sheets_name: list[str] = wb_daily_report.sheetnames
    sheets_name = tuple(full_sheets_name[int(start_day - 1) : int(end_day - 1)])

    full_data = []

    for day in sheets_name:
        # باز کردن شیت روز های کل ماه
        ws_daily_report = wb_daily_report[day]

        data = [
            ws_daily_report["F9"].value,  # تاریخ
            ws_daily_report["F9"].value,  # ذرات زیر ۱۰
            ws_daily_report["F10"].value,  # ذرات ۰ تا 5
            ws_daily_report["G10"].value,  # ذرات ۵ تا ۱۰
            ws_daily_report["H9"].value,  # ذرات ۱۰ تا ۶۰
            ws_daily_report["K9"].value,  # ذرات بالای ۶۰
            ws_daily_report["A9"].value,  # تناژ
            ws_daily_report["B9"].value,  # تعداد کامیون
        ]
        # اگر حتی یکی از مقادیر عددی نبود
        if any(not isinstance(day_test, (int, float)) for day_test in data):
            click.echo(message=f"❌ Data of day {day} is empty - skipping...")
            continue

        # دریافت اطلاعات مربوط به سل های مورد بررسی
        clean_data = [
            str(data[0]).strip().split("/"),
            Decimal(str(data[1])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات زیر ۱۰
            Decimal(str(data[2])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۰ تا 5
            Decimal(str(data[3])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۵ تا ۱۰
            Decimal(str(data[4])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۱۰ تا ۶۰
            Decimal(str(data[5])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات بالای ۶۰
            Decimal(str(data[6])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # تناژ
            Decimal(str(data[7])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # تعداد کامیون
        ]

        full_data.append(tuple(clean_data))

    wb_daily_report.close()

    return full_data


if __name__ == "__main__":
    data = cli.main(["load-data"], standalone_mode=False)
    print(data)
