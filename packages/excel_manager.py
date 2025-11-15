from openpyxl import Workbook, load_workbook
from openpyxl.utils import cell
from typing import List, Any

class ExcelCellManager:
    """
    Aclass for reading and writing cells and ranges in Excel files using openpyxl.
    """
    
    def __init__(self, filename: str):
        """
        Initialize the Excel manager with a filename.
        
        Args:
            filename (str): Path to the Excel file
        """
        self.filename = filename
        self.workbook = None
        self.active_ws = None
    
    def load_workbook(self, create_if_missing: bool = False) -> None:
        """
        Load an existing workbook or create a new one.
        
        Args:
            create_if_missing (bool): Create new workbook if file doesn't exist
        """
        try:
            self.workbook = load_workbook(self.filename)
            self.active_ws = self.workbook.active
        except FileNotFoundError:
            if create_if_missing:
                self.workbook = Workbook()
                self.active_ws = self.workbook.active
                self.save()
            else:
                raise FileNotFoundError(f"File {self.filename} not found")
    
    def set_active_sheet(self, sheet_name: str) -> None:
        """
        Set the active worksheet.
        
        Args:
            sheet_name (str): Name of the worksheet
        """
        if self.workbook is None:
            self.load_workbook()
        
        if sheet_name in self.workbook.sheetnames:
            self.active_ws = self.workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found")
    
    def read_cells(self, cell_references: List[str]) -> List[Any]:
        """
        Read values from a list of cell references.
        
        Args:
            cell_references (List[str]): List of cell references (e.g., ['A1', 'B2', 'C3'])
        
        Returns:
            List[Any]: List of cell values in the same order as input references
        
        Example:
            >>> manager.read_cells(['A1', 'B2', 'C3'])
            ['Value1', 'Value2', 'Value3']
        """
        if self.workbook is None:
            self.load_workbook()
        
        values = []
        for cell_ref in cell_references:
            try:
                cell_addr = cell.coordinate_from_string(cell_ref)
                values.append(self.active_ws[cell_ref].value)
            except Exception as e:
                raise ValueError(f"Invalid cell reference: {cell_ref} - {str(e)}")
        
        return values
    
    def write_cells(self, cell_references: List[str], values: List[Any]) -> None:
        """
        Write values to a list of cell references.
        
        Args:
            cell_references (List[str]): List of cell references (e.g., ['A1', 'B2', 'C3'])
            values (List[Any]): List of values to write
        
        Example:
            >>> manager.write_cells(['A1', 'B2', 'C3'], ['Hello', 42, 3.14])
        """
        if self.workbook is None:
            self.load_workbook()
        
        if len(cell_references) != len(values):
            raise ValueError("Number of cell references must match number of values")
        
        for cell_ref, value in zip(cell_references, values):
            try:
                self.active_ws[cell_ref] = value
            except Exception as e:
                raise ValueError(f"Invalid cell reference: {cell_ref} - {str(e)}")
        
        self.save()
    
    def read_range(self, start_cell: str, end_cell: str) -> List[List[Any]]:
        """
        Read a rectangular range of cells.
        
        Args:
            start_cell (str): Top-left cell of the range (e.g., 'A1')
            end_cell (str): Bottom-right cell of the range (e.g., 'C3')
        
        Returns:
            List[List[Any]]: 2D list of values (rows x columns)
        
        Example:
            >>> manager.read_range('A1', 'C3')
            [['A1', 'B1', 'C1'],
             ['A2', 'B2', 'C2'],
             ['A3', 'B3', 'C3']]
        """
        if self.workbook is None:
            self.load_workbook()
        
        try:
            start_col, start_row = cell.coordinate_from_string(start_cell)
            end_col, end_row = cell.coordinate_from_string(end_cell)
            
            # Convert column letters to indices
            start_col_idx = cell.column_index_from_string(start_col)
            end_col_idx = cell.column_index_from_string(end_col)
            
            range_data = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col_idx, end_col_idx + 1):
                    col_letter = cell.get_column_letter(col)
                    cell_ref = f"{col_letter}{row}"
                    row_data.append(self.active_ws[cell_ref].value)
                range_data.append(row_data)
            
            return range_data
            
        except Exception as e:
            raise ValueError(f"Invalid range: {start_cell}:{end_cell} - {str(e)}")
    
    def write_range(self, start_cell: str, data: List[List[Any]]) -> None:
        """
        Write a 2D list of values to a range starting from the specified cell.
        
        Args:
            start_cell (str): Top-left cell where data will be written
            data (List[List[Any]]): 2D list of values to write
        
        Example:
            >>> data = [['Name', 'Age'], ['Alice', 25], ['Bob', 30]]
            >>> manager.write_range('A1', data)
        """
        if self.workbook is None:
            self.load_workbook()
        
        if not data or not data[0]:
            raise ValueError("Data cannot be empty")
        
        try:
            start_col, start_row = cell.coordinate_from_string(start_cell)
            start_col_idx = cell.column_index_from_string(start_col)
            
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    current_col = cell.get_column_letter(start_col_idx + col_idx)
                    current_row = start_row + row_idx
                    cell_ref = f"{current_col}{current_row}"
                    self.active_ws[cell_ref] = value
            
            self.save()
            
        except Exception as e:
            raise ValueError(f"Invalid start cell or data: {start_cell} - {str(e)}")
        
    
    def save(self, filename: str = None) -> None:
        """Save the workbook to file."""
        if self.workbook is None:
            raise ValueError("No workbook loaded")
        
        save_name = filename or self.filename
        self.workbook.save(save_name)
    
    def close(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close()

