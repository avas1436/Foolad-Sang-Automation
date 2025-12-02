from decimal import ROUND_HALF_UP, Decimal

import click
from openpyxl import load_workbook


@click.command()
@click.option(
    '--start_day', prompt='Enter start day', default=1, type=int, show_default=True
)
@click.option(
    '--end_day', prompt='Enter end day', default=31, type=int, show_default=True
)
@click.option(
    '--file_path', prompt='Enter file path', default='daily.xlsx', show_default=True
)
def check_daily(file_path: str, start_day: int, end_day: int):
    # اگر پارامترها داده نشده باشند، مقدار پیش‌فرض بده
    if start_day < 0:
        start_day = 0
    if end_day > 30:
        end_day = 30
    # باز کردن و خواندن اکسل گزارش روزانه در حالت خواندن برای سرعت بیشتر
    try:
        wb_daily_report = load_workbook(
            filename=rf"{file_path}", data_only=True, read_only=True
        )
    except Exception as e:
        click.secho(message=f"There is a problem with excel file: {e}", fg="red")

    sheets_name: list[str] = wb_daily_report.sheetnames

    # یک حلقه که روی تک تک روزهای ماه حرکت کرده و بررسی
    #  می‌کند آیا محاسبات صحیح بوده یا نه
    for day in range(start_day - 1, end_day - 1):

        # باز کردن شیت روز های کل ماه
        ws_daily_report = wb_daily_report[sheets_name[day]]

        granulation_data_test = [
            ws_daily_report["F9"].value,  # ذرات زیر ۱۰
            ws_daily_report["F10"].value,  # ذرات ۰ تا 5
            ws_daily_report["G10"].value,  # ذرات ۵ تا ۱۰
            ws_daily_report["H9"].value,  # ذرات ۱۰ تا ۶۰
            ws_daily_report["K9"].value,  # ذرات بالای ۶۰
            ws_daily_report["A9"].value,  # تناژ
            ws_daily_report["B9"].value,  # تعداد کامیون
        ]
        # اگر حتی یکی از مقادیر عددی نبود
        if any(
            not isinstance(day_test, (int, float)) for day_test in granulation_data_test
        ):
            click.echo(f"❌ Data of day {day+1} is empty - skipping...")
            continue

        # دریافت اطلاعات مربوط به سل های مورد بررسی
        granulation_data = [
            Decimal(str(granulation_data_test[0])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات زیر ۱۰
            Decimal(str(granulation_data_test[1])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۰ تا 5
            Decimal(str(granulation_data_test[2])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۵ تا ۱۰
            Decimal(str(granulation_data_test[3])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات ۱۰ تا ۶۰
            Decimal(str(granulation_data_test[4])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # ذرات بالای ۶۰
            Decimal(str(granulation_data_test[5])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # تناژ
            Decimal(str(granulation_data_test[6])).quantize(
                Decimal("0.1"), rounding=ROUND_HALF_UP
            ),  # تعداد کامیون
        ]

        # شروط مورد نظر برای بررسی صحت اطلاعات
        if (granulation_data[1] + granulation_data[2] != granulation_data[0]) or (
            granulation_data[0] + granulation_data[3] + granulation_data[4]
            != Decimal("100.000")
        ):
            click.echo(message=f"⚠️  Day {day+1}: Granulation calculation error")
        if granulation_data[1] + granulation_data[2] != granulation_data[0]:
            click.echo(
                message=f"   Sum of 0-5 ({granulation_data[1]}) and 5-10 ({granulation_data[2]})"
            )
            click.echo(message=f"   does not equal below 10 ({granulation_data[0]})")
            click.echo(message="   ─────────────────────────")
        if granulation_data[0] + granulation_data[3] + granulation_data[4] != Decimal(
            "100.000"
        ):
            click.echo(message=f"   Below 10: {granulation_data[0]}")
            click.echo(message=f"   10-60:    {granulation_data[3]}")
            click.echo(message=f"   Above 60: {granulation_data[4]}")
            click.echo(message=f"   does not equal 100%")
            click.echo(message="   ─────────────────────────")
        if granulation_data[6] and granulation_data[6] != 0:
            avg_tonnage = granulation_data[5] / granulation_data[6]
            if avg_tonnage < Decimal("24.2") or avg_tonnage > Decimal("26.8"):
                click.echo(
                    message=f"⚠️  Day {day+1}: Tonnage or number of truck is not correct!"
                )
                click.echo(message="   ─────────────────────────")

    wb_daily_report.close()


# اجرای کاملا ساده با کلیک که تقریبا همه ارور هندلینگ ها را خودش انجام میده
if __name__ == '__main__':
    check_daily()
